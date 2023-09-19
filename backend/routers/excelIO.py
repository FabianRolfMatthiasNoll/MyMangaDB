import base64
import io
import re
import time
import zipfile
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session
from io import BytesIO

import crud.author as authorManager
import crud.genre as genreManager
import crud.volume as volumeManager
import crud.manga as mangaManager

from database import get_db
from schema import Volume
from models import Manga as DBManga

router = APIRouter(prefix="/excel", tags=["ExcelInOut"])

##########################################  Import   ###########################################################

# TODO: If the title exists already ask the User if he wants to import the new one, keep the old one or keep both


def validate_manga_fields(row):
    """Check if required manga fields are present and return errors if any."""
    (
        title,
        description,
        genres_str,
        authors_str,
        _,
        _,
        reading_status,
        collection_status,
    ) = row
    errors = []
    field_names = [
        "title",
        "description",
        "genres_str",
        "authors_str",
        "reading_status",
        "collection_status",
    ]
    values = [
        title,
        description,
        genres_str,
        authors_str,
        reading_status,
        collection_status,
    ]
    for field, value in zip(field_names, values):
        if not value:
            errors.append(f"Missing {field} in Manga {title}")
    return errors


def process_genres(db, genres_str, manga_model):
    """Process and relate genres to the given manga model."""
    genre_names = (
        [genre.strip() for genre in genres_str.split(",")] if genres_str else []
    )
    for genre_name in genre_names:
        genre = genreManager.get_genre(db, genre_name) or genreManager.create_genre(
            db, genre_name
        )
        genreManager.create_relation(db, genre.id, manga_model.id)


def process_authors(db, authors_str, manga_model):
    """Process authors and relate them to the given manga model."""
    authors_list = (
        [author.strip() for author in authors_str.split(",")] if authors_str else []
    )
    for author_name in authors_list:
        author = authorManager.get_author(
            db, author_name
        ) or authorManager.create_author(db, author_name)
        authorManager.create_relation(db, author.id, manga_model.id)


@router.post("/import")
async def import_mangas_from_excel(
    file: UploadFile = File(...), db: Session = Depends(get_db)
):
    # Initialize empty dictionary for cover images
    cover_images = {}

    # Buffer to hold the Excel file data
    excel_buffer = None

    # Extract data from the zip archive
    with zipfile.ZipFile(io.BytesIO(await file.read()), "r") as zf:
        for file_info in zf.filelist:
            if file_info.filename.startswith("cover_images/"):
                with zf.open(file_info.filename) as img_file:
                    img_data = img_file.read()
                    cover_images[file_info.filename.split("/")[-1]] = base64.b64encode(
                        img_data
                    ).decode()
            elif file_info.filename.endswith(
                ".xlsx"
            ):  # Assuming Excel files have .xlsx extension
                with zf.open(file_info.filename) as excel_file:
                    excel_buffer = excel_file.read()

    # Load the Excel workbook
    wb = load_workbook(filename=io.BytesIO(excel_buffer))
    ws = wb.active
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Validate required fields before processing
        row_errors = validate_manga_fields(row)
        if row_errors:
            errors.extend(row_errors)

        title, _, _, _, total_volumes, specific_volumes_str, _, _ = row

        # Create main Manga entry
        manga_model = DBManga(
            title=title,
            description=row[1],
            total_volumes=total_volumes,
            reading_status=row[6],
            collection_status=row[7],
        )

        # Associate cover image
        manga_cover_name = f"{title}.jpg"
        if manga_cover_name in cover_images:
            manga_model.cover_image = cover_images[manga_cover_name]

        db.add(manga_model)
        db.commit()
        db.refresh(manga_model)

        # Process genres, authors, roles
        process_genres(db, row[2], manga_model)
        process_authors(db, row[3], manga_model)

        # Process Volumes
        volume_nums = (
            [int(vol.strip()) for vol in specific_volumes_str.split(",")]
            if specific_volumes_str
            else []
        )
        for volume_num in volume_nums:
            volume = Volume(
                id=0, volume_num=volume_num, manga_id=manga_model.id, cover_image=""
            )

            # Associate cover image for volume
            volume_cover_name = f"{title}_volume_{volume_num}.jpg"
            if volume_cover_name in cover_images:
                volume.cover_image = cover_images[volume_cover_name]

            volumeManager.create_volume(db, volume)

    if errors:
        return {"message": errors}
    return {"message": "Mangas imported successfully"}


##########################################  Export   ###########################################################


def create_headers(ws):
    """Add headers to the worksheet."""
    ws.append(
        [
            "Title",
            "Description",
            "Genres",
            "Authors",
            "Total Volumes",
            "Specific Volumes",
            "Reading Status",
            "Collection Status",
        ]
    )


def get_manga_data(manga):
    """Retrieve manga related data to be used in the worksheet."""
    genres = ", ".join([genre.name for genre in manga.genres])
    authors = ", ".join([author.name for author in manga.authors])
    specific_volumes = ", ".join([str(volume.volume_num) for volume in manga.volumes])
    return (
        manga.title,
        manga.description,
        genres,
        authors,
        len(manga.volumes),
        specific_volumes,
        manga.reading_status.value,
        manga.collection_status.value,
    )


@router.get("/export")
async def export_mangas_to_excel(db: Session = Depends(get_db)):
    # Fetch and process manga data
    db_mangas = db.query(DBManga).all()
    mangas = [mangaManager.create_manga_model(db, db_manga) for db_manga in db_mangas]

    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Mangas"

    # Add headers
    create_headers(ws)

    # Populate worksheet with manga data
    for manga in mangas:
        ws.append(get_manga_data(manga))

    # Save workbook to a buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)

    # Create a zip buffer
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("MyMangaDB_library.xlsx", excel_buffer.getvalue())

        # Save cover images
        for manga in mangas:
            if manga.cover_image:
                # Decode the base64 string and write as image
                img_data = base64.b64decode(manga.cover_image)
                zf.writestr(f"cover_images/{manga.title}.jpg", img_data)

            for volume in manga.volumes:
                if volume.cover_image:
                    img_data = base64.b64decode(volume.cover_image)
                    zf.writestr(
                        f"cover_images/{manga.title}_volume_{volume.volume_num}.jpg",
                        img_data,
                    )

    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=MyMangaDB_library.zip"},
    )
