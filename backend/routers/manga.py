from typing import List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from io import BytesIO
from schema import Author, Genre

import crud.author as authorManager
import crud.genre as genreManager
import crud.volume as volumeManager
import crud.manga as mangaManager

from database import get_db
from schema import Manga, Volume
from models import Manga as DBManga
from models import Volume as DBVolume

router = APIRouter(prefix="/manga", tags=["Manga"])

# TODO: Try to fetch cover images for mangas
# TODO: Get all Genres
# TODO: Get all Authors
# TODO: Get all Roles

# TODO: Will be for now withold until it is certain that these features are needed:
# Remove Author endpoint
# Remove Genre endpoint
############################################################################################################


@router.get("/")
def get_all_mangas(db: Session = Depends(get_db)) -> List[Manga]:
    db_mangas = db.query(DBManga).all()
    mangas = []
    for db_manga in db_mangas:
        mangas.append(mangaManager.create_manga_model(db, db_manga))
    return mangas


@router.get("/id/{manga_id}")
def get_manga_by_id(manga_id: int, db: Session = Depends(get_db)) -> Manga:
    db_manga = db.query(DBManga).filter(DBManga.id == manga_id).one_or_none()
    if db_manga is None:
        raise HTTPException(status_code=404, detail="Manga id not found")
    return mangaManager.create_manga_model(db, db_manga)


@router.get("/title/{manga_title}")
def get_manga_by_title(manga_title: str, db: Session = Depends(get_db)) -> Manga:
    manga = mangaManager.get_manga_by_title(db, manga_title)
    if manga is None:
        raise HTTPException(status_code=404, detail="Manga title not found")
    return mangaManager.create_manga_model(db, manga)


@router.get("/genre/{genre_name}")
def get_mangas_by_genre(genre_name: str, db: Session = Depends(get_db)) -> List[Manga]:
    genre = genreManager.get_genre(db, genre_name)
    if genre is None:
        raise HTTPException(status_code=404, detail="Genre not found")
    relations = genreManager.get_relations_by_genre_id(db, genre.id)
    return mangaManager.get_mangas_by_relations(db, relations)


@router.get("/author/{author_name}")
def get_mangas_by_author(
    author_name: str, db: Session = Depends(get_db)
) -> List[Manga]:
    author = authorManager.get_author(db, author_name)
    if author is None:
        raise HTTPException(status_code=404, detail="Author not found")
    relations = authorManager.get_relations_by_author_id(db, author.id)
    return mangaManager.get_mangas_by_relations(db, relations)


############################################################################################################


@router.post("/volume")
def add_volume(volume: Volume, db: Session = Depends(get_db)) -> Manga:
    existing_volumes = volumeManager.get_volumes_by_manga_id(db, volume.manga_id)
    for existing_volume in existing_volumes:
        if existing_volume.volume_num == volume.volume_num:
            raise HTTPException(status_code=404, detail="Volume already existing")

    volumeManager.create_volume(db, volume)

    return get_manga_by_id(volume.manga_id, db)


# TODO: Change to modify volume.
@router.put("/volume/cover")
def add_cover_to_volume(volume: Volume, db: Session = Depends(get_db)):
    db_manga = db.query(DBVolume).filter(DBVolume.id == volume.id).one_or_none()

    db_manga.cover_image = volume.cover_image

    db.commit()
    db.refresh(db_manga)


@router.delete("/volume")
def remove_volume(volume_id: int, db: Session = Depends(get_db)):
    volumeManager.delete_volume(db, volume_id)


############################################################################################################


@router.post("/")
def create_manga(manga: Manga, db: Session = Depends(get_db)) -> Manga:
    return mangaManager.create_manga(db, manga)


@router.put("/update_manga")
def update_manga(manga: Manga, db: Session = Depends(get_db)) -> Manga:
    return mangaManager.update_manga(db, manga)


@router.delete("/remove")
def remove_manga(manga_id: int, db: Session = Depends(get_db)):
    mangaManager.remove_manga(db, manga_id)


############################################################################################################


@router.get("/genre")
def get_all_genre_names(db: Session = Depends(get_db)) -> List[str]:
    return genreManager.get_all_genre_names(db)


@router.get("/authors")
def get_all_author_names(db: Session = Depends(get_db)) -> List[str]:
    return authorManager.get_all_author_names(db)


@router.get("/authors/roles")
def get_all_role_names(db: Session = Depends(get_db)) -> List[str]:
    return authorManager.get_all_role_names(db)


############################################################################################################


@router.get("/export")
async def export_mangas_to_excel(db: Session = Depends(get_db)):
    db_mangas = db.query(DBManga).all()
    mangas = [mangaManager.create_manga_model(db, db_manga) for db_manga in db_mangas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Mangas"

    # Add headers
    ws.append(
        [
            "Title",
            "Description",
            "Genres",
            "Authors(Roles)",
            "Total Volumes",
            "Specific Volumes",
        ]
    )

    # Add data
    for manga in mangas:
        genres = ", ".join([genre.name for genre in manga.genres])
        authors_roles = ", ".join(
            [f"{author.name}({author.role})" for author in manga.authors]
        )
        specific_volumes = ", ".join(
            [str(volume.volume_num) for volume in manga.volumes]
        )

        ws.append(
            [
                manga.title,
                manga.description,
                genres,
                authors_roles,
                len(manga.volumes),
                specific_volumes,
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mangas.xlsx"},
    )


@router.post("/import")
async def import_mangas_from_excel(
    file: UploadFile = File(...), db: Session = Depends(get_db)
):
    wb = load_workbook(filename=file.file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            title,
            description,
            genres_str,
            authors_roles_str,
            total_volumes,
            specific_volumes_str,
        ) = row

        manga_model = DBManga()
        manga_model.title = title
        manga_model.description = description
        manga_model.totalVolumes = total_volumes
        db.add(manga_model)
        db.commit()
        db.refresh(manga_model)

        # Handle Genres
        genre_names = (
            [genre.strip() for genre in genres_str.split(",")] if genres_str else []
        )
        for genre_name in genre_names:
            genre = genreManager.get_genre(db, genre_name)
            if not genre:
                genre = genreManager.create_genre(db, genre_name)
            genreManager.create_relation(db, genre.id, manga_model.id)

        authors_list = (
            [author_role.strip() for author_role in authors_roles_str.split(",")]
            if authors_roles_str
            else []
        )
        #TODO: Watch out for cases like missing role etc, after hookup to frontend
        # add error messages etc.
        for author_role_str in authors_list:
            parts = author_role_str.split("(")
            author_name = parts[0].strip()
            role = parts[1].replace(")", "").strip() if len(parts) > 1 else None
            result_author = authorManager.get_author(db, author_name)
            if result_author is None:
                result_author = authorManager.create_author(db, author_name)
            result_role = authorManager.get_role(db, role)
            if result_role is None:
                result_role = authorManager.create_role(db, role)
            authorManager.create_relation(
                db, result_author.id, manga_model.id, result_role.id
            )

        # Handle Volumes
        volume_nums = (
            [int(vol.strip()) for vol in specific_volumes_str.split(",")]
            if specific_volumes_str
            else []
        )
        for volume_num in volume_nums:
            volume = Volume(
                id=0,
                volume_num=volume_num,
                manga_id=manga_model.id,
                cover_image="",
            )
            volumeManager.create_volume(db, volume)

    return {"message": "Mangas imported successfully"}
